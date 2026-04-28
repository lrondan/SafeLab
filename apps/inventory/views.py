from urllib import request

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Campus, Laboratory, Equipment, Reagent, Glassware, Component, SafeMaterial, OtherItem, ProcessTrainer
from .forms import EquipmentForm, ReagentForm, GlasswareForm, ComponentForm, SafeMaterialForm, OtherItemForm, ProcessTrainerForm
import pubchempy as pcp
from django.core.paginator import Paginator

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


@login_required
def home(request):
    """Main page: list of all campuses."""
    campuses = Campus.objects.all()
    return render(request, 'inventory/home.html', {'campuses': campuses})


@login_required
def campus_detail(request, campus_id):
    """Show all laboratories in a campus."""
    campus = get_object_or_404(Campus, id=campus_id)
    laboratories = campus.laboratories.all()
    return render(request, 'inventory/campus_detail.html', {
        'campus': campus,
        'laboratories': laboratories
    })


@login_required
def lab_detail(request, lab_id):
    """Laboratory detail with Equipment and Reagents (Bootstrap tabs)."""
    lab = get_object_or_404(Laboratory, id=lab_id)

    # Equipment
    equipments = lab.equipments.all().order_by('name')
    equipment_status = request.GET.get('equipment_status')
    if equipment_status:
        equipments = equipments.filter(status=equipment_status)

    # Reagents
    reagents = lab.reagents.all().order_by('common_name')
    reagent_status = request.GET.get('reagent_status')
    if reagent_status:
        reagents = reagents.filter(status=reagent_status)

    #glassware
    glasswares = lab.glasswares.all().order_by('name')
    glassware_status = request.GET.get('glassware_status')
    if glassware_status:
        glasswares = glasswares.filter(status=glassware_status)

    #components
    components = lab.components.all().order_by('name')
    component_status = request.GET.get('component_status')
    if component_status:
        components = components.filter(status=component_status)

    #safe_materials
    safe_materials = lab.safematerials.all().order_by('name')
    safe_materials_status = request.GET.get('safe_materials_status')
    if safe_materials_status:
        safe_materials = safe_materials.filter(status=safe_materials_status)

    #other_items
    other_items = lab.otheritems.all().order_by('name')
    other_items_status = request.GET.get('other_items_status')
    if other_items_status:
        other_items = other_items.filter(status=other_items_status)

    process_trainers = lab.processtrainers.all().order_by('model')
    process_trainers_status = request.GET.get('process_trainers_status')
    if process_trainers_status:
        process_trainers = process_trainers.filter(status=process_trainers_status)

    def paginate(qs, param):
        return Paginator(qs, 10).get_page(request.GET.get(param))

    equipments = paginate(equipments, 'equipments_page')
    reagents = paginate(reagents, 'reagents_page')
    glasswares = paginate(glasswares, 'glasswares_page')
    components = paginate(components, 'components_page')
    safe_materials = paginate(safe_materials, 'safe_materials_page')
    process_trainer = paginate(process_trainers, 'process_trainer_page')
    other_items = paginate(other_items, 'other_items_page')


    return render(request, 'inventory/lab_detail.html', {
        'lab': lab,
        'equipments': equipments,
        'reagents': reagents,
        'glasswares': glasswares,
        'components': components,
        'safe_materials': safe_materials,
        'process_trainers': process_trainers,
        'other_items': other_items,
        'equipment_statuses': Equipment.STATUSES,
        'reagent_statuses': Reagent.STATUSES,
        'glassware_statuses': Glassware.STATUSES,
        'component_statuses': Component.STATUSES,
        'process_trainer_statuses': ProcessTrainer.STATUSES,
    })


@login_required
def equipment_create(request, lab_id):
    """Create new equipment."""
    lab = get_object_or_404(Laboratory, id=lab_id)
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            equipment = form.save(commit=False)
            equipment.laboratory = lab
            equipment.save()
            return redirect('lab_detail', lab_id=lab.id)
    else:
        form = EquipmentForm()

    return render(request, 'inventory/equipment_form.html', {
        'form': form,
        'lab': lab,
        'title': 'New Equipment'
    })

@login_required
def equipment_update(request, equipment_id):
    """Update equipment (e.g. mark as Broken)."""
    equipment = get_object_or_404(Equipment, id=equipment_id)
    
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            form.save()
            return redirect('lab_detail', lab_id=equipment.laboratory.id)
    else:
        form = EquipmentForm(instance=equipment)

    return render(request, 'inventory/equipment_form.html', {
        'form': form,
        'equipment': equipment,
        'lab': equipment.laboratory,    
        'title': 'Edit Equipment'
    })


@login_required
def equipment_delete(request, equipment_id):
    """Delete equipment."""
    equipment = get_object_or_404(Equipment, id=equipment_id)
    lab_id = equipment.laboratory.id
    if request.method == 'POST':
        equipment.delete()
        return redirect('lab_detail', lab_id=lab_id)

    return render(request, 'inventory/equipment_confirm_delete.html', {
        'equipment': equipment,
        'lab': equipment.laboratory
    })

@login_required
def component_create(request, lab_id):
    """Create new component."""
    lab = get_object_or_404(Laboratory, id=lab_id)
    if request.method == 'POST':
        form = ComponentForm(request.POST)
        if form.is_valid():
            component = form.save(commit=False)
            component.laboratory = lab
            component.save()
            return redirect('lab_detail', lab_id=lab.id)
    else:
        form = ComponentForm()

    return render(request, 'inventory/component_create.html', {
        'form': form,
        'lab': lab,
        'title': 'New Component'
    })

@login_required
def component_update(request, component_id):
    """Update component."""
    component = get_object_or_404(Component, id=component_id)
    
    if request.method == 'POST':
        form = ComponentForm(request.POST, instance=component)
        if form.is_valid():
            form.save()
            return redirect('lab_detail', lab_id=component.laboratory.id)
    else:
        form = ComponentForm(instance=component)

    return render(request, 'inventory/component_update.html', {
        'form': form,
        'component': component,
        'lab': component.laboratory,    
        'title': 'Edit Component'
    })

@login_required
def component_delete(request, component_id):
    """Delete component."""
    component = get_object_or_404(Component, id=component_id)
    lab_id = component.laboratory.id
    if request.method == 'POST':
        component.delete()
        return redirect('lab_detail', lab_id=lab_id)

    return render(request, 'inventory/component_confirm_delete.html', {
        'component': component,
        'lab': component.laboratory
    })

@login_required
def reagent_create(request, lab_id):
    """Create reagent with PubChem auto-fill (MSDS link)."""
    lab = get_object_or_404(Laboratory, id=lab_id)
    message = ""
    form = None

    if request.method == 'POST':
        if 'search_pubchem' in request.POST and request.POST.get('search_pubchem'):
            term = request.POST['search_pubchem'].strip()
            try:
                compounds = pcp.get_compounds(term, 'name')
                if compounds:
                    c = compounds[0]
                    initial_data = {
                        'common_name': term,
                        'iupac_name': c.iupac_name or '',
                        'formula': c.molecular_formula or '',
                        'molecular_weight': float(c.molecular_weight) if c.molecular_weight else None,
                        'pubchem_cid': c.cid,
                    }
                    form = ReagentForm(initial=initial_data)
                    message = f"✅ Data loaded from PubChem (CID: {c.cid}). Review and save."
                else:
                    form = ReagentForm(request.POST)
                    message = "⚠️ No compound found in PubChem."
            except Exception as e:
                form = ReagentForm(request.POST)
                message = f"⚠️ PubChem error: {str(e)}"
        else:
            form = ReagentForm(request.POST)
            if form.is_valid():
                reagent = form.save(commit=False)
                reagent.laboratory = lab
                reagent.save()
                return redirect('lab_detail', lab_id=lab.id)
    else:
        form = ReagentForm()

    return render(request, 'inventory/reagent_form.html', {
        'form': form,
        'lab': lab,
        'message': message
    })


@login_required
def reagent_update(request, reagent_id):
    """Update reagent."""
    reagent = get_object_or_404(Reagent, id=reagent_id)
    if request.method == 'POST':
        form = ReagentForm(request.POST, instance=reagent)
        if form.is_valid():
            form.save()
            return redirect('lab_detail', lab_id=reagent.laboratory.id)
    else:
        form = ReagentForm(instance=reagent)

    return render(request, 'inventory/reagent_form.html', {
        'form': form,
        'reagent': reagent,
        'lab': reagent.laboratory
    })

@login_required
def reagent_delete(request, reagent_id):
    """Delete reagent."""
    reagent = get_object_or_404(Reagent, id=reagent_id)
    lab_id = reagent.laboratory.id
    if request.method == 'POST':
        reagent.delete()
        return redirect('lab_detail', lab_id=lab_id)

    return render(request, 'inventory/reagent_confirm_delete.html', {
        'reagent': reagent,
        'lab': reagent.laboratory
    })

@login_required
def glassware_create(request, lab_id):
    """Create new glassware."""
    lab = get_object_or_404(Laboratory, id=lab_id)
    if request.method == 'POST':
        form = GlasswareForm(request.POST)
        if form.is_valid():
            glassware = form.save(commit=False)
            glassware.laboratory = lab
            glassware.save()
            return redirect('lab_detail', lab_id=lab.id)
    else:
        form = GlasswareForm()

    return render(request, 'inventory/glassware_form.html', {
        'form': form,
        'lab': lab,
        'title': 'New Glassware'
    })


@login_required
def glassware_update(request, glassware_id):
    """Update glassware."""
    glassware = get_object_or_404(Glassware, id=glassware_id)
    if request.method == 'POST':
        form = GlasswareForm(request.POST, instance=glassware)
        if form.is_valid():
            form.save()
            return redirect('lab_detail', lab_id=glassware.laboratory.id)
    else:
        form = GlasswareForm(instance=glassware)

    return render(request, 'inventory/glassware_form.html', {
        'form': form,
        'glassware': glassware,
        'lab': glassware.laboratory,
        'title': 'Edit Glassware'
    })

@login_required
def glassware_delete(request, glassware_id):
    """Delete glassware."""
    glassware = get_object_or_404(Glassware, id=glassware_id)
    lab_id = glassware.laboratory.id
    if request.method == 'POST':
        glassware.delete()
        return redirect('lab_detail', lab_id=lab_id)

    return render(request, 'inventory/glassware_confirm_delete.html', {
        'glassware': glassware,
        'lab': glassware.laboratory
    })

@login_required
def safe_material_create(request, lab_id):
    """Create new safe material."""
    lab = get_object_or_404(Laboratory, id=lab_id)
    if request.method == 'POST':
        form = SafeMaterialForm(request.POST)
        if form.is_valid():
            safe_material = form.save(commit=False)
            safe_material.laboratory = lab
            safe_material.save()
            return redirect('lab_detail', lab_id=lab.id)
    else:
        form = SafeMaterialForm()

    return render(request, 'inventory/safe_material_form.html', {
        'form': form,
        'lab': lab,
        'title': 'New Safe Material'
    })

@login_required
def safe_material_update(request, safe_material_id):
    """Update safe material."""
    safe_material = get_object_or_404(SafeMaterial, id=safe_material_id)
    if request.method == 'POST':
        form = SafeMaterialForm(request.POST, instance=safe_material)
        if form.is_valid():
            form.save()
            return redirect('lab_detail', lab_id=safe_material.laboratory.id)
    else:
        form = SafeMaterialForm(instance=safe_material)

    return render(request, 'inventory/safe_material_form.html', {
        'form': form,
        'safe_material': safe_material,
        'lab': safe_material.laboratory,
        'title': 'Edit Safe Material'
    })

@login_required
def safe_material_delete(request, safe_material_id):
    """Delete safe material."""
    safe_material = get_object_or_404(SafeMaterial, id=safe_material_id)
    lab_id = safe_material.laboratory.id
    if request.method == 'POST':
        safe_material.delete()
        return redirect('lab_detail', lab_id=lab_id)

    return render(request, 'inventory/safe_material_confirm_delete.html', {
        'safe_material': safe_material,
        'lab': safe_material.laboratory
    })


@login_required
def process_trainer_update(request, process_trainer_id):
    process_trainer = get_object_or_404(ProcessTrainer, id=process_trainer_id)
    if request.method == 'POST':
        form = ProcessTrainerForm(request.POST, instance=process_trainer)
        if form.is_valid():
            form.save()
            return redirect('lab_detail', lab_id=process_trainer.laboratory.id)
    else:
        form = ProcessTrainerForm(instance=process_trainer)

    return render(request, 'inventory/process_trainer_form.html', {
        'form': form,
        'process_trainer': process_trainer,
        'lab': process_trainer.laboratory,
        'title': 'Edit Safe Process Trainner'
    })

@login_required
def process_trainer_create(request, lab_id):
    lab = get_object_or_404(Laboratory, id=lab_id)
    if request.method == 'POST':
        form = ProcessTrainerForm(request.POST)
        if form.is_valid():
            process_trainer = form.save(commit=False)
            process_trainer.laboratory = lab
            process_trainer.save()
            return redirect('lab_detail', lab_id=lab.id)
    else:
        form = ProcessTrainerForm()

    return render(request, 'inventory/process_trainer_form.html', {
        'form': form,
        'lab': lab,
        'title': 'New Process Trainer'
    })

@login_required
def process_trainer_delete(request, process_trainer_id):
    """Delete process trainer."""
    process_trainer = get_object_or_404(ProcessTrainer, id=process_trainer_id)
    lab_id = process_trainer.laboratory.id
    if request.method == 'POST':
        process_trainer.delete()
        return redirect('lab_detail', lab_id=lab_id)

    return render(request, 'inventory/process_trainer_delete.html', {
        'process_trainer': process_trainer,
        'lab': process_trainer.laboratory
    })

@login_required
def other_item_create(request, lab_id):
    """Create new other item."""
    lab = get_object_or_404(Laboratory, id=lab_id)
    if request.method == 'POST':
        form = OtherItemForm(request.POST)
        if form.is_valid():
            other_item = form.save(commit=False)
            other_item.laboratory = lab
            other_item.save()
            return redirect('lab_detail', lab_id=lab.id)
    else:
        form = OtherItemForm()

    return render(request, 'inventory/other_item_form.html', {
        'form': form,
        'lab': lab,
        'title': 'New Other Item'
    })

@login_required
def other_item_update(request, other_item_id):
    """Update other item."""
    other_item = get_object_or_404(OtherItem, id=other_item_id)
    if request.method == 'POST':
        form = OtherItemForm(request.POST, instance=other_item)
        if form.is_valid():
            form.save()
            return redirect('lab_detail', lab_id=other_item.laboratory.id)
    else:
        form = OtherItemForm(instance=other_item)

    return render(request, 'inventory/other_item_form.html', {
        'form': form,
        'other_item': other_item,
        'lab': other_item.laboratory,
        'title': 'Edit Other Item'
    })

@login_required
def other_item_delete(request, other_item_id):
    """Delete other item."""
    other_item = get_object_or_404(OtherItem, id=other_item_id)
    lab_id = other_item.laboratory.id
    if request.method == 'POST':
        other_item.delete()
        return redirect('lab_detail', lab_id=lab_id)

    return render(request, 'inventory/other_item_confirm_delete.html', {
        'other_item': other_item,
        'lab': other_item.laboratory
    })

@login_required
def export_lab_to_excel(request, lab_id):
    lab = get_object_or_404(Laboratory, id=lab_id)

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    # ====================== EQUIPMENT SHEET ======================
    ws_equip = wb.create_sheet("Equipment")
    headers_equip = ['Name', 'Model', 'Serial Number', 'Quantity', 'Status', 'Description', 'Updated At']
    
    # Encabezados bonitos
    for col, header in enumerate(headers_equip, 1):
        cell = ws_equip.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")

    equipments = lab.equipments.all().order_by('name')
    for row_idx, eq in enumerate(equipments, 2):
        ws_equip.cell(row=row_idx, column=1, value=eq.name)
        ws_equip.cell(row=row_idx, column=2, value=eq.model_name or "")
        ws_equip.cell(row=row_idx, column=3, value=eq.serial_number)
        ws_equip.cell(row=row_idx, column=4, value=eq.quantity)
        ws_equip.cell(row=row_idx, column=5, value=eq.get_status_display())
        ws_equip.cell(row=row_idx, column=6, value=eq.description or "")
        ws_equip.cell(row=row_idx, column=7, value=eq.updated_at.strftime("%Y-%m-%d %H:%M"))

    # Auto-ajustar columnas
    for col in ws_equip.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_equip.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    # Título grande
    ws_equip.insert_rows(1)
    title_cell = ws_equip['A1']
    title_cell.value = f"Lab Inventory Report - {lab.name} ({lab.campus.name})"
    title_cell.font = Font(bold=True, size=14)
    ws_equip.merge_cells('A1:G1')

# ====================== GLASSWARE SHEET ======================
    ws_glass = wb.create_sheet("Glassware")
    headers_glass = ['Name', 'Description', 'Volume', 'Quantity', 'Status', 'Notes', 'Updated At']

    # Encabezados bonitos
    for col, header in enumerate(headers_glass, 1):
        cell = ws_glass.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")

    glasswares = lab.glasswares.all().order_by('name')
    for row_idx, gl in enumerate(glasswares, 2):
        ws_glass.cell(row=row_idx, column=1, value=gl.name)
        ws_glass.cell(row=row_idx, column=2, value=gl.description or "")
        ws_glass.cell(row=row_idx, column=3, value=gl.volume)
        ws_glass.cell(row=row_idx, column=4, value=gl.quantity)
        ws_glass.cell(row=row_idx, column=5, value=gl.get_status_display())
        ws_glass.cell(row=row_idx, column=6, value=gl.notes or "")
        ws_glass.cell(row=row_idx, column=7, value=gl.date_updated.strftime("%Y-%m-%d %H:%M"))

    # Auto-ajustar columnas
    for col in ws_glass.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_glass.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    # Título grande
    ws_glass.insert_rows(1)
    title_cell = ws_glass['A1']
    title_cell.value = f"Lab Inventory Report - {lab.name} ({lab.campus.name})"
    title_cell.font = Font(bold=True, size=14)
    ws_glass.merge_cells('A1:G1')

# ====================== COMPONENT SHEET ======================
    ws_comp = wb.create_sheet("Components")
    headers_comp = ['Name', 'Description', 'Quantity', 'Status', 'Notes', 'Updated At']

    # Encabezados bonitos
    for col, header in enumerate(headers_comp, 1):
        cell = ws_comp.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")

    components = lab.components.all().order_by('name')
    for row_idx, comp in enumerate(components, 2):
        ws_comp.cell(row=row_idx, column=1, value=comp.name)
        ws_comp.cell(row=row_idx, column=2, value=comp.description or "")
        ws_comp.cell(row=row_idx, column=3, value=comp.quantity)
        ws_comp.cell(row=row_idx, column=4, value=comp.get_status_display())
        ws_comp.cell(row=row_idx, column=5, value=comp.notes or "")
        ws_comp.cell(row=row_idx, column=6, value=comp.date_updated.strftime("%Y-%m-%d %H:%M"))

    # Auto-ajustar columnas
    for col in ws_comp.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_comp.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    # Título grande
    ws_comp.insert_rows(1)
    title_cell = ws_comp['A1']
    title_cell.value = f"Lab Inventory Report - {lab.name} ({lab.campus.name})"
    title_cell.font = Font(bold=True, size=14)
    ws_comp.merge_cells('A1:G1')

    # ====================== REAGENTS SHEET ======================
    ws_reag = wb.create_sheet("Reagents")
    headers_reag = ['Common Name', 'Formula', 'Quantity', 'Unit', 'Status', 'CAS Number', 'PubChem CID', 'Safety Notes', 'Updated At']
    
    for col, header in enumerate(headers_reag, 1):
        cell = ws_reag.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")

    reagents = lab.reagents.all().order_by('common_name')
    for row_idx, r in enumerate(reagents, 2):
        ws_reag.cell(row=row_idx, column=1, value=r.common_name)
        ws_reag.cell(row=row_idx, column=2, value=r.formula or "")
        ws_reag.cell(row=row_idx, column=3, value=r.quantity)
        ws_reag.cell(row=row_idx, column=4, value=r.unit)
        ws_reag.cell(row=row_idx, column=5, value=r.get_status_display())
        ws_reag.cell(row=row_idx, column=6, value=r.cas_number or "")
        ws_reag.cell(row=row_idx, column=7, value=r.pubchem_cid or "")
        ws_reag.cell(row=row_idx, column=8, value=r.safety_notes or "")
        ws_reag.cell(row=row_idx, column=9, value=r.updated_at.strftime("%Y-%m-%d %H:%M"))

    # Auto-ajustar
    for col in ws_reag.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_reag.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    # Título
    ws_reag.insert_rows(1)
    title_cell = ws_reag['A1']
    title_cell.value = f"Lab Inventory Report - {lab.name} ({lab.campus.name})"
    title_cell.font = Font(bold=True, size=14)
    ws_reag.merge_cells('A1:I1')

    # ====================== SAFE MATERIALS SHEET ======================
    ws_safe = wb.create_sheet("Safe Materials")
    headers_safe = ['Name', 'Description', 'Quantity', 'Status', 'Notes', 'Updated At']

    for col, header in enumerate(headers_safe, 1):
        cell = ws_safe.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    safe_materials = lab.safematerials.all().order_by('name')
    for row_idx, sm in enumerate(safe_materials, 2):
        ws_safe.cell(row=row_idx, column=1, value=sm.name)
        ws_safe.cell(row=row_idx, column=2, value=sm.description or "")
        ws_safe.cell(row=row_idx, column=3, value=sm.quantity)
        ws_safe.cell(row=row_idx, column=4, value=sm.get_status_display())
        ws_safe.cell(row=row_idx, column=5, value=sm.notes or "")
        ws_safe.cell(row=row_idx, column=6, value=sm.date_updated.strftime("%Y-%m-%d %H:%M"))

    # Auto-ajustar
    for col in ws_safe.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_safe.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
    # Título
    ws_safe.insert_rows(1)
    title_cell = ws_safe['A1']
    title_cell.value = f"Lab Inventory Report - {lab.name} ({lab.campus.name})"
    title_cell.font = Font(bold=True, size=14)
    ws_safe.merge_cells('A1:F1')

    # ====================== PROCESS TRAINERS SHEET ======================
    ws_process = wb.create_sheet("Process Trainers")
    headers_process = ['Model', 'Serial Number', 'Quantity', 'Description', 'Status', 'Updated At']

    for col, header in enumerate(headers_process, 1):
        cell = ws_process.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    process_trainers = lab.processtrainers.all().order_by('model')
    for row_idx, pt in enumerate(process_trainers, 2):
        ws_process.cell(row=row_idx, column=1, value=pt.model)
        ws_process.cell(row=row_idx, column=2, value=pt.serial_number)
        ws_process.cell(row=row_idx, column=3, value=pt.quantity)
        ws_process.cell(row=row_idx, column=5, value=pt.description or "")
        ws_process.cell(row=row_idx, column=4, value=pt.get_status_display())
        ws_process.cell(row=row_idx, column=6, value=pt.date_updated.strftime("%Y-%m-%d %H:%M"))

    # Auto-ajustar
    for col in ws_process.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_process.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
    # Título
    ws_process.insert_rows(1)
    title_cell = ws_process['A1']
    title_cell.value = f"Lab Inventory Report - {lab.name} ({lab.campus.name})"
    title_cell.font = Font(bold=True, size=14)
    ws_process.merge_cells('A1:F1')


    # ====================== OTHER ITEMS SHEET ======================
    ws_other = wb.create_sheet("Other Items")
    headers_other = ['Name', 'Description', 'Quantity', 'Status', 'Notes', 'Updated At']
    for col, header in enumerate(headers_other, 1):
        cell = ws_other.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    other_items = lab.otheritems.all().order_by('name')
    for row_idx, oi in enumerate(other_items, 2):   
        ws_other.cell(row=row_idx, column=1, value=oi.name)
        ws_other.cell(row=row_idx, column=2, value=oi.description or "")
        ws_other.cell(row=row_idx, column=3, value=oi.quantity)
        ws_other.cell(row=row_idx, column=4, value=oi.get_status_display())
        ws_other.cell(row=row_idx, column=5, value=oi.notes or "")
        ws_other.cell(row=row_idx, column=6, value=oi.date_updated.strftime("%Y-%m-%d %H:%M"))
    # Auto-ajustar
    for col in ws_other.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_other.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
    # Título
    ws_other.insert_rows(1)
    title_cell = ws_other['A1']
    title_cell.value = f"Lab Inventory Report - {lab.name} ({lab.campus.name})"
    title_cell.font = Font(bold=True, size=14)
    ws_other.merge_cells('A1:F1')

    # =================== Respuesta ===================
    filename = f"Lab_Inventory_{lab.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response