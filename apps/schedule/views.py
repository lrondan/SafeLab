import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import LabSession, SchedulePeriod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


# ── main grid ────────────────────────────────────────────────────

@login_required
def schedule(request):
    session = SchedulePeriod.objects.filter(active=True).first()
    lab_session = []
    if session:
        lab_session = session.period_sessions.all()

    return render(request, 'schedules/schedule.html', {
        'session':       session,
        'count':         LabSession.objects.all().order_by('day'),
        'lab_sessions':  lab_session
    })

# ── Excel Report ─────────────────────────────────────────────────
@login_required
def export_practicals_to_excel(request, session_id):
    # Obtener el periodo
    session = get_object_or_404(SchedulePeriod, id=session_id)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Sessions"

    # =================== TÍTULO ===================
    ws['A1'] = f"Lab Sessions Report - {session.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:I1')

    # =================== ENCABEZADOS ===================
    headers = [
        'Period', 'Laboratory', 'Day', 'Professor',
        'Start Time', 'End Time', 'Activity',
        'Students', 'Status'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1f4e79",
            end_color="1f4e79",
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # =================== DATOS ===================
    practicals = session.period_sessions.all().order_by('day')

    for row_idx, eq in enumerate(practicals, start=3):
        ws.cell(row=row_idx, column=1, value=session.name)
        ws.cell(row=row_idx, column=2, value=str(eq.laboratory))
        ws.cell(row=row_idx, column=3, value=eq.day.strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=4, value=eq.professor)
        ws.cell(row=row_idx, column=5, value=eq.start_time.strftime('%H:%M'))
        ws.cell(row=row_idx, column=6, value=eq.end_time.strftime('%H:%M'))
        ws.cell(row=row_idx, column=7, value=eq.activity)
        ws.cell(row=row_idx, column=8, value=eq.student_count)
        ws.cell(
            row=row_idx,
            column=9,
            value="Completed" if eq.practice_complete else "Pending"
        )

    # =================== AUTO AJUSTE ===================
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # =================== RESPUESTA ===================
    filename = f"Lab_Sessions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response